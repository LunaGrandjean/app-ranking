import re
import io
import datetime as dt
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule


st.set_page_config(page_title="Analyse CDF + Ranking", layout="wide")

# =========================================================
# PAGE 1 — ANALYSE CDF
# =========================================================
EXPECTED_COLS = [
    "Rank", "Phase", "about", "date", "competition", "Location", "Description", "Total",
    "score_finale", "mouches", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "S10", "S11", "S12",
    "P1", "P2", "P3", "P4", "P5", "Épreuve", "source_pdf"
]
SERIES_COLS = [f"S{i}" for i in range(1, 13)]
NUM_COLS = ["Rank", "Total", "score_finale"] + SERIES_COLS + [f"P{i}" for i in range(1, 6)]


@st.cache_data(show_spinner=False)
def load_data_from_uploads(uploaded_files) -> pd.DataFrame:
    dfs = []
    for f in uploaded_files:
        df = pd.read_csv(f)
        df["file_name"] = f.name
        dfs.append(df)
    if not dfs:
        return pd.DataFrame()
    return pd.concat(dfs, ignore_index=True)


@st.cache_data(show_spinner=False)
def load_data_from_folder(folder_path: str) -> pd.DataFrame:
    data_dir = Path(folder_path)
    csv_files = sorted(data_dir.glob("*.csv"))
    dfs = []
    for f in csv_files:
        df = pd.read_csv(f)
        df["file_name"] = f.name
        dfs.append(df)
    if not dfs:
        return pd.DataFrame()
    return pd.concat(dfs, ignore_index=True)



def normalize_athlete_name(name: str) -> str:
    name = str(name).strip().upper()
    name = re.sub(r"\s+", " ", name)
    return name



def parse_event_info(pdf_name: str) -> pd.Series:
    name = str(pdf_name).lower()

    discipline = "50m" if "50" in name else "10m"

    if "women" in name:
        sexe = "Women"
    elif "men" in name:
        sexe = "Men"
    else:
        sexe = "unknown"

    if "cadet" in name:
        categorie_age = "Cadet"
    elif "junior" in name:
        categorie_age = "Junior"
    else:
        categorie_age = "Senior"

    return pd.Series([discipline, sexe, categorie_age])


@st.cache_data(show_spinner=False)
def prepare_data(raw: pd.DataFrame) -> pd.DataFrame:
    if raw.empty:
        return raw.copy()

    df = raw.copy()

    for c in EXPECTED_COLS:
        if c not in df.columns:
            df[c] = np.nan

    for c in ["about", "Phase", "source_pdf", "Description", "Épreuve"]:
        df[c] = df[c].astype(str).str.strip()

    for c in NUM_COLS:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["year"] = df["date"].dt.year

    df[["discipline", "sexe", "categorie_age"]] = df["source_pdf"].apply(parse_event_info)
    df["nb_series"] = df[SERIES_COLS].notna().sum(axis=1)
    df["Total_normalise_6"] = df["Total"]

    mask_10m = (df["discipline"] == "10m") & (df["nb_series"] > 0)
    df.loc[mask_10m, "Total_normalise_6"] = (
        df.loc[mask_10m, "Total"] * 6 / df.loc[mask_10m, "nb_series"]
    )

    df["athlete"] = df["about"].apply(normalize_athlete_name)

    df = df[
        (df["Rank"].notna())
        & (df["Rank"] > 0)
        & (df["Total"].notna())
        & (df["Total"] > 0)
        & (df["nb_series"] > 0)
    ].copy()

    return df



def annual_stats(data: pd.DataFrame, discipline: str, sexe: str, score_col: str = "Total") -> pd.DataFrame:
    sub = data[
        (data["discipline"] == discipline)
        & (data["sexe"] == sexe)
    ].copy()

    sub = sub.dropna(subset=["Rank", score_col, "year"])
    sub["Rank"] = pd.to_numeric(sub["Rank"], errors="coerce")
    sub = sub.dropna(subset=["Rank"])
    sub["Rank"] = sub["Rank"].astype(int)

    out = []
    for year, g in sub.groupby("year"):
        g = g.sort_values("Rank")
        out.append(
            {
                "year": year,
                "first_score": g.loc[g["Rank"] == 1, score_col].mean(),
                "podium_mean": g[g["Rank"] <= 3][score_col].mean(),
                "top10_mean": g[g["Rank"] <= 10][score_col].mean(),
                "rank10_score": g.loc[g["Rank"] == 10, score_col].mean(),
            }
        )

    return pd.DataFrame(out).sort_values("year") if out else pd.DataFrame()



def plot_annual_scores(df: pd.DataFrame, discipline: str, sexe: str, categorie_age: str, score_col: str = "Total"):
    dff = df[
        (df["discipline"] == discipline)
        & (df["sexe"] == sexe)
        & (df["categorie_age"] == categorie_age)
    ].copy()

    stats = annual_stats(dff, discipline, sexe, score_col=score_col)
    if stats.empty:
        return None

    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=stats["year"],
            y=stats["rank10_score"],
            mode="lines",
            line=dict(width=0),
            showlegend=False,
            hoverinfo="skip",
        )
    )
    fig.add_trace(
        go.Scatter(
            x=stats["year"],
            y=stats["first_score"],
            mode="lines",
            fill="tonexty",
            name="Couloir 1er-10e",
            hovertemplate="Année=%{x}<br>1er=%{y:.2f}<extra></extra>",
        )
    )
    fig.add_trace(
        go.Scatter(
            x=stats["year"],
            y=stats["first_score"],
            mode="lines+markers",
            name="1er",
            hovertemplate="Année=%{x}<br>1er=%{y:.2f}<extra></extra>",
        )
    )
    fig.add_trace(
        go.Scatter(
            x=stats["year"],
            y=stats["podium_mean"],
            mode="lines+markers",
            name="Moyenne podium",
            hovertemplate="Année=%{x}<br>Podium=%{y:.2f}<extra></extra>",
        )
    )
    fig.add_trace(
        go.Scatter(
            x=stats["year"],
            y=stats["top10_mean"],
            mode="lines+markers",
            name="Moyenne top 10",
            hovertemplate="Année=%{x}<br>Top10=%{y:.2f}<extra></extra>",
        )
    )
    fig.update_layout(
        title=f"Évolution annuelle des scores — {discipline} / {sexe} / {categorie_age}",
        xaxis_title="Année",
        yaxis_title="Score",
        template="plotly_white",
        hovermode="x unified",
    )
    return fig



def plot_athlete_score(data: pd.DataFrame, athlete_name: str, discipline: str, sexe: Optional[str] = None):
    athlete_norm = normalize_athlete_name(athlete_name)
    sub = data[(data["athlete"] == athlete_norm) & (data["discipline"] == discipline)].copy()
    if sexe is not None:
        sub = sub[sub["sexe"] == sexe]
    if sub.empty:
        return None

    sub = sub.sort_values(["year", "categorie_age"])
    fig = px.line(
        sub,
        x="year",
        y="Total",
        color="categorie_age",
        markers=True,
        hover_data=["Rank", "discipline", "sexe", "source_pdf"],
        title=f"Évolution du score — {athlete_name} ({discipline})",
    )
    fig.update_layout(template="plotly_white", xaxis_title="Année", yaxis_title="Score")
    return fig



def plot_athlete_rank(data: pd.DataFrame, athlete_name: str, discipline: str, sexe: Optional[str] = None):
    athlete_norm = normalize_athlete_name(athlete_name)
    sub = data[(data["athlete"] == athlete_norm) & (data["discipline"] == discipline)].copy()
    if sexe is not None:
        sub = sub[sub["sexe"] == sexe]
    if sub.empty:
        return None

    sub = sub.sort_values(["year", "categorie_age"])
    fig = px.line(
        sub,
        x="year",
        y="Rank",
        color="categorie_age",
        markers=True,
        hover_data=["Total", "discipline", "sexe", "source_pdf"],
        title=f"Évolution du rank — {athlete_name} ({discipline})",
    )
    fig.update_layout(template="plotly_white", xaxis_title="Année", yaxis_title="Rank")
    fig.update_yaxes(autorange="reversed")
    return fig



def athlete_summary(data: pd.DataFrame, athlete_name: str) -> pd.DataFrame:
    athlete_norm = normalize_athlete_name(athlete_name)
    sub = data[data["athlete"] == athlete_norm].copy()
    if sub.empty:
        return pd.DataFrame()
    return sub[["about", "year", "discipline", "sexe", "categorie_age", "Rank", "Total", "source_pdf"]].sort_values(["discipline", "year"])



def top_athletes(data: pd.DataFrame, categorie_age: str, discipline: str, n: int = 15) -> pd.DataFrame:
    sub = (
        data[(data["categorie_age"] == categorie_age) & (data["discipline"] == discipline)]
        .dropna(subset=["Rank"])
        .sort_values(["Rank", "Total"], ascending=[True, False])
        .drop_duplicates("athlete")
        .head(n)
    )
    return sub[["athlete", "year", "Rank", "Total", "sexe", "source_pdf"]]


# =========================================================
# PAGE 2 — RANKING APP
# =========================================================
EXCEL_ENGINE = "openpyxl"
DENOM = 100.0


def excel_date_to_date(x):
    if pd.isna(x):
        return pd.NaT
    if isinstance(x, (int, float, np.integer, np.floating)):
        return (pd.Timestamp("1899-12-30") + pd.to_timedelta(int(x), unit="D")).date()
    if isinstance(x, pd.Timestamp):
        return x.date()

    s = str(x).strip()
    if not s:
        return pd.NaT

    dtv = pd.to_datetime(s, errors="coerce", format="%Y-%m-%d %H:%M:%S")
    if pd.isna(dtv):
        dtv = pd.to_datetime(s, errors="coerce", dayfirst=True)

    return dtv.date() if not pd.isna(dtv) else pd.NaT



def squish(x) -> str:
    if pd.isna(x):
        return ""
    return re.sub(r"\s+", " ", str(x)).strip()



def round_existing_columns(df: pd.DataFrame, cols: list, decimals: int = 1) -> pd.DataFrame:
    df = df.copy()
    for col in cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").round(decimals)
    return df



def within_last_12_months(d, date_ref):
    if pd.isna(d) or pd.isna(date_ref):
        return False
    age_days = (date_ref - d).days
    return 0 <= age_days < 365



def extract_sheet_long(raw: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    raw = raw.copy()

    header_row = None
    for r in range(min(40, raw.shape[0])):
        c0 = squish(raw.iat[r, 0]).lower()
        c1 = squish(raw.iat[r, 1]).lower()
        if c0 == "cat" and c1 == "name":
            header_row = r
            break

    if header_row is None:
        return pd.DataFrame(columns=["Athlete", "Category", "Competition", "Date", "Score", "Rank", "Sheet"])

    athlete_start_row = header_row + 1
    comp_row = max(0, header_row - 2)

    score_cols = []
    for c in range(raw.shape[1]):
        if squish(raw.iat[header_row, c]).lower() == "score":
            score_cols.append(c)

    if not score_cols:
        return pd.DataFrame(columns=["Athlete", "Category", "Competition", "Date", "Score", "Rank", "Sheet"])

    first_comp_col = score_cols[0]
    step = (score_cols[1] - score_cols[0]) if len(score_cols) >= 2 else 3

    athletes = raw.iloc[athlete_start_row:, 1].apply(squish).replace({"": np.nan})
    categories = raw.iloc[athlete_start_row:, 0].apply(squish).replace({"": np.nan})

    parts = []
    for col in range(first_comp_col, raw.shape[1] - 1, step):
        comp_name = squish(raw.iat[comp_row, col])
        if comp_name == "":
            continue

        comp_date_raw = raw.iat[comp_row, col + 1] if col + 1 < raw.shape[1] else pd.NaT
        comp_date = excel_date_to_date(comp_date_raw)

        score = pd.to_numeric(raw.iloc[athlete_start_row:, col], errors="coerce")
        rank = pd.to_numeric(raw.iloc[athlete_start_row:, col + 1], errors="coerce")

        df_part = pd.DataFrame({
            "Athlete": athletes,
            "Category": categories,
            "Competition": comp_name,
            "Date": comp_date,
            "Score": score,
            "Rank": rank,
            "Sheet": sheet_name,
        })

        df_part = df_part[df_part["Athlete"].notna()]
        df_part = df_part[~(df_part["Score"].isna() & df_part["Rank"].isna())]
        parts.append(df_part)

    if not parts:
        return pd.DataFrame(columns=["Athlete", "Category", "Competition", "Date", "Score", "Rank", "Sheet"])

    return pd.concat(parts, ignore_index=True)



def extract_workbook(file_bytes: bytes) -> pd.DataFrame:
    xf = pd.ExcelFile(io.BytesIO(file_bytes), engine=EXCEL_ENGINE)
    dfs = []
    for sheet in xf.sheet_names:
        try:
            raw = pd.read_excel(
                io.BytesIO(file_bytes),
                sheet_name=sheet,
                header=None,
                engine=EXCEL_ENGINE,
                dtype=object,
            )
            dfs.append(extract_sheet_long(raw, sheet))
        except Exception:
            pass

    if not dfs:
        return pd.DataFrame(columns=["Athlete", "Category", "Competition", "Date", "Score", "Rank", "Sheet"])
    return pd.concat(dfs, ignore_index=True)



def add_comp_coeff(df: pd.DataFrame, coeff_table: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    ct = coeff_table.copy()
    ct.columns = [c.strip().upper() for c in ct.columns]

    if "MATCH" not in ct.columns or "COEFF" not in ct.columns:
        raise ValueError("La table coeff doit contenir les colonnes MATCH et COEFF")

    ct["MATCH"] = ct["MATCH"].astype(str).str.strip()
    ct["COEFF"] = pd.to_numeric(ct["COEFF"], errors="coerce")
    ct = ct.dropna(subset=["MATCH", "COEFF"])

    rules = [(rf"\b{re.escape(m)}\b", float(v)) for m, v in zip(ct["MATCH"], ct["COEFF"])]
    comp = df["Competition"].fillna("").astype(str)

    def coeff_one(s: str) -> float:
        for pat, val in rules:
            if re.search(pat, s):
                return val
        return np.nan

    df["Coeff"] = comp.apply(coeff_one)
    return df



def fix_names(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    repl = {r"^BRIAN Julie$": "BRIAND Julie"}
    df["Athlete"] = df["Athlete"].replace(repl, regex=True)
    df["Athlete"] = df["Athlete"].replace({
        "BRACKMAN Anceline": "BRACKMAN Ancéline",
        "CADET Madisson": "CADET Madison",
        "BLATEAUX Félix": "BLATEAU Félix",
        "VEDEL Thimothée": "VEDEL Timothée",
        "TENERIFE Alaisia": "TENERIFE Alaisa",
        "BONNET Marcellin": "BONNET Marcelin",
        "CHOLET Mathis": "CHOLET Matis",
    })
    df["Athlete"] = df["Athlete"].str.replace(r"IUNG Antoine\*+", "IUNG Antoine", regex=True)
    return df



def parse_scale_from_excel(file_bytes: bytes, sheet_name: str = None) -> pd.DataFrame:
    bio = io.BytesIO(file_bytes)
    xf = pd.ExcelFile(bio, engine=EXCEL_ENGINE)
    use_sheet = sheet_name or xf.sheet_names[0]
    raw = pd.read_excel(bio, sheet_name=use_sheet, header=None, engine=EXCEL_ENGINE)

    wanted = {
        "50m Women", "50m Men", "50m Junior Women", "50m Junior Men",
        "10m Women", "10m Men", "10m Junior Women", "10m Junior Men",
    }

    headers = {}
    for r in range(raw.shape[0]):
        for c in range(raw.shape[1]):
            v = raw.iat[r, c]
            if isinstance(v, str):
                t = v.strip()
                if t in wanted:
                    headers[t] = (r, c)

    rows = []
    for title, (r0, c0) in headers.items():
        rr = r0 + 1
        while rr < raw.shape[0]:
            score = raw.iat[rr, c0]
            pts = raw.iat[rr, c0 + 1] if c0 + 1 < raw.shape[1] else np.nan
            if pd.isna(score) and pd.isna(pts):
                break
            score_num = pd.to_numeric(score, errors="coerce")
            pts_num = pd.to_numeric(pts, errors="coerce")
            if not pd.isna(score_num) and not pd.isna(pts_num):
                rows.append({"ScaleKey": title, "MinScore": float(score_num), "Points": float(pts_num)})
            rr += 1

    scale = pd.DataFrame(rows)
    if scale.empty:
        return scale

    return scale.sort_values(["ScaleKey", "MinScore"], ascending=[True, False]).reset_index(drop=True)



def score_to_points(score: float, scale_df: pd.DataFrame, key: str) -> float:
    if pd.isna(score):
        return np.nan

    sub = scale_df[scale_df["ScaleKey"] == key].copy()
    if sub.empty:
        return np.nan

    sub["MinScore"] = pd.to_numeric(sub["MinScore"], errors="coerce")
    sub["Points"] = pd.to_numeric(sub["Points"], errors="coerce")
    sub = sub.dropna(subset=["MinScore", "Points"]).sort_values("MinScore", ascending=False)

    if sub.empty:
        return np.nan

    min_threshold = sub["MinScore"].min()
    if score < min_threshold:
        return 0.0

    for _, row in sub.iterrows():
        if score >= row["MinScore"]:
            return float(row["Points"])

    return 0.0



def derive_sex_distance(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    sheet = df["Sheet"].fillna("").astype(str)

    df["Sexe"] = np.where(sheet.str.contains("DAMES", case=False, na=False), "F", "M")
    df["Distance"] = None
    df.loc[sheet.str.contains("10m", case=False, na=False), "Distance"] = "10m"
    df.loc[sheet.str.contains("50m", case=False, na=False), "Distance"] = "50m"
    return df



def scale_key(distance: str, sexe: str, category: str) -> str:
    if pd.isna(distance) or pd.isna(sexe):
        return None

    cat = str(category).strip().upper()
    is_junior = cat.startswith("J")
    women = str(sexe).upper() == "F"

    if distance == "50m":
        if is_junior and women:
            return "50m Junior Women"
        if is_junior and not women:
            return "50m Junior Men"
        if (not is_junior) and women:
            return "50m Women"
        return "50m Men"

    if distance == "10m":
        if is_junior and women:
            return "10m Junior Women"
        if is_junior and not women:
            return "10m Junior Men"
        if (not is_junior) and women:
            return "10m Women"
        return "10m Men"

    return None



def default_final_points():
    return pd.DataFrame({
        "Rank": [1, 2, 3, 4, 5, 6, 7, 8],
        "FinalPoints": [30, 28, 26, 18, 16, 14, 12, 10],
    })



def add_date_coeff(df: pd.DataFrame, date_ref: dt.date) -> pd.DataFrame:
    df = df.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date

    def f(d):
        if pd.isna(d):
            return np.nan
        age_days = (date_ref - d).days
        if age_days <= 122:
            return 1.0
        if age_days > 365:
            return 0.0
        k = 3.0
        x = (365 - age_days) / (365 - 122)
        return (1 - np.exp(-k * x)) / (1 - np.exp(-k))

    df["date_coeff"] = df["Date"].apply(f)
    return df



def make_final_table(df: pd.DataFrame, date_ref: dt.date) -> pd.DataFrame:
    df = df.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date

    def weighted_mean(vals, w):
        vals = np.array(vals, dtype=float)
        w = np.array(w, dtype=float)
        mask = ~np.isnan(vals) & ~np.isnan(w)
        if mask.sum() == 0:
            return np.nan
        denom = w[mask].sum()
        if denom == 0:
            return np.nan
        return float((vals[mask] * w[mask]).sum() / denom)

    rows = []
    for (ath, dist, sexe, cat), g in df.groupby(["Athlete", "Distance", "Sexe", "Category"], dropna=False):
        g = g.copy()
        g["is_12m"] = g["Date"].apply(lambda d: within_last_12_months(d, date_ref))
        g_valid = g[g["is_12m"]].copy()

        n12 = int((~g_valid["Score"].isna()).sum())
        av12 = float(g_valid["Score"].mean()) if g_valid["Score"].notna().any() else np.nan
        hp12 = float(g_valid["Score"].max()) if g_valid["Score"].notna().any() else np.nan

        pctJ = weighted_mean(g_valid["%J"], g_valid["Coeff"])
        pctS = weighted_mean(g_valid["%S"], g_valid["Coeff"])

        weights_time = g["Coeff"] * g["date_coeff"]
        pctJ_t = weighted_mean(g["%J"], weights_time)
        pctS_t = weighted_mean(g["%S"], weights_time)

        finale_avg = g_valid["Finale_score"].dropna().mean() if g_valid["Finale_score"].notna().any() else np.nan

        row = {
            "Athlete": ath,
            "Distance": dist,
            "%J": pctJ,
            "%J compet": pctJ_t,
            "Finale J": finale_avg,
            "%S": pctS,
            "%S compet": pctS_t,
            "Finale S": finale_avg,
            "Moyenne": av12,
            "HP": hp12,
            "Nombre de compet": n12,
            "Sexe": sexe,
            "Catégorie": cat,
        }

        if n12 < 5:
            for kcol in ["%J", "%J compet", "%S", "%S compet"]:
                row[kcol] = 0

        if str(cat).strip().upper() == "S":
            row["%J"] = np.nan
            row["%J compet"] = np.nan
            row["Finale J"] = np.nan

        rows.append(row)

    final_tbl = pd.DataFrame(rows)
    ordered_cols = [
        "Athlete", "Distance", "%J", "%J compet", "Finale J",
        "%S", "%S compet", "Finale S", "Moyenne", "HP",
        "Nombre de compet", "Sexe", "Catégorie",
    ]
    final_tbl = final_tbl[ordered_cols]
    final_tbl = round_existing_columns(
        final_tbl,
        ["%J", "%J compet", "Finale J", "%S", "%S compet", "Finale S", "Moyenne", "HP"],
        1,
    )

    if "Nombre de compet" in final_tbl.columns:
        final_tbl["Nombre de compet"] = pd.to_numeric(final_tbl["Nombre de compet"], errors="coerce").fillna(0).astype(int)

    return final_tbl


# =========================================================
# RENDER FUNCTIONS
# =========================================================
def render_page_cdf():
    st.title("Analyse CDF — version Streamlit")
    st.caption("Page 1 : analyse des CSV extraits du notebook CDF.ipynb")

    with st.sidebar:
        st.header("Chargement des données CSV")
        source_mode = st.radio(
            "Source des CSV",
            ["Dossier local csv", "Upload manuel"],
            index=0,
            key="cdf_source_mode",
        )

        if source_mode == "Dossier local csv":
            folder = st.text_input("Chemin du dossier CSV", value="csv", key="cdf_folder")
            raw = load_data_from_folder(folder)
        else:
            uploaded_files = st.file_uploader(
                "Importer un ou plusieurs CSV",
                type=["csv"],
                accept_multiple_files=True,
                key="cdf_uploads",
            )
            raw = load_data_from_uploads(uploaded_files) if uploaded_files else pd.DataFrame()

    df = prepare_data(raw) if not raw.empty else pd.DataFrame()

    if raw.empty:
        st.info("Aucun fichier CSV chargé. Place tes fichiers dans un dossier `csv/` à côté du script, ou importe-les manuellement dans la barre latérale.")
        return

    if df.empty:
        st.warning("Les fichiers ont été chargés, mais aucune ligne exploitable n'a été trouvée après nettoyage.")
        return

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Lignes brutes", len(raw))
    col2.metric("Lignes analysables", len(df))
    col3.metric("Athlètes", df["athlete"].nunique())
    col4.metric("Années", int(df["year"].nunique()) if df["year"].notna().any() else 0)

    with st.expander("Aperçu des données préparées"):
        st.dataframe(df.head(50), use_container_width=True)

    tab1, tab2, tab3, tab4 = st.tabs([
        "Scores annuels",
        "Analyse athlète",
        "Top athlètes",
        "Données",
    ])

    with tab1:
        st.subheader("Évolution annuelle des scores")
        c1, c2, c3 = st.columns(3)
        discipline = c1.selectbox("Discipline", sorted(df["discipline"].dropna().unique()), key="annual_discipline")
        sexe = c2.selectbox("Sexe", sorted(df["sexe"].dropna().unique()), key="annual_sexe")
        categorie = c3.selectbox("Catégorie", sorted(df["categorie_age"].dropna().unique()), key="annual_cat")

        score_col = "Total_normalise_6" if discipline == "10m" else "Total"
        fig = plot_annual_scores(df, discipline, sexe, categorie, score_col=score_col)
        if fig is None:
            st.warning("Pas de données pour cette combinaison.")
        else:
            st.plotly_chart(fig, use_container_width=True)

    with tab2:
        st.subheader("Analyse par athlète")
        athletes = sorted(df["athlete"].dropna().unique())
        athlete = st.selectbox("Athlète", athletes, key="cdf_athlete")
        sex_filter_options = ["Tous"] + sorted(df[df["athlete"] == athlete]["sexe"].dropna().unique().tolist())
        sex_filter = st.selectbox("Filtre sexe", sex_filter_options, key="cdf_sex_filter")
        sex_filter_value = None if sex_filter == "Tous" else sex_filter

        summary = athlete_summary(df, athlete)
        if sex_filter_value is not None and not summary.empty:
            summary = summary[summary["sexe"] == sex_filter_value]

        if summary.empty:
            st.warning("Athlète non trouvé.")
        else:
            st.dataframe(summary, use_container_width=True)
            available_disciplines = sorted(summary["discipline"].dropna().unique())
            for disc in available_disciplines:
                score_fig = plot_athlete_score(df, athlete, disc, sexe=sex_filter_value)
                rank_fig = plot_athlete_rank(df, athlete, disc, sexe=sex_filter_value)
                if score_fig is not None:
                    st.plotly_chart(score_fig, use_container_width=True)
                if rank_fig is not None:
                    st.plotly_chart(rank_fig, use_container_width=True)

    with tab3:
        st.subheader("Top athlètes")
        c1, c2, c3 = st.columns(3)
        top_cat = c1.selectbox("Catégorie âge", sorted(df["categorie_age"].dropna().unique()), key="top_cat")
        top_disc = c2.selectbox("Discipline", sorted(df["discipline"].dropna().unique()), key="top_disc")
        top_n = c3.slider("Nombre d'athlètes", min_value=5, max_value=30, value=15, key="top_n")

        st.dataframe(top_athletes(df, top_cat, top_disc, top_n), use_container_width=True)

    with tab4:
        st.subheader("Exploration libre")
        years = sorted([int(y) for y in df["year"].dropna().unique()])
        selected_years = st.multiselect("Années", years, default=years, key="cdf_years")
        selected_sexes = st.multiselect("Sexes", sorted(df["sexe"].dropna().unique()), default=sorted(df["sexe"].dropna().unique()), key="cdf_sexes")
        selected_cats = st.multiselect("Catégories", sorted(df["categorie_age"].dropna().unique()), default=sorted(df["categorie_age"].dropna().unique()), key="cdf_cats")
        selected_disc = st.multiselect("Disciplines", sorted(df["discipline"].dropna().unique()), default=sorted(df["discipline"].dropna().unique()), key="cdf_disc")

        filtered = df[
            df["year"].isin(selected_years)
            & df["sexe"].isin(selected_sexes)
            & df["categorie_age"].isin(selected_cats)
            & df["discipline"].isin(selected_disc)
        ].copy()

        st.dataframe(filtered, use_container_width=True)
        st.download_button(
            "Télécharger les données filtrées en CSV",
            data=filtered.to_csv(index=False).encode("utf-8"),
            file_name="cdf_filtre.csv",
            mime="text/csv",
            key="cdf_download",
        )



def render_page_ranking():
    st.title("Ranking App (Excel → paramètres → tableau final)")
    st.caption("Page 2 : calcul du tableau final à partir du fichier résultats FFTir")

    if "final_tbl" not in st.session_state:
        st.session_state["final_tbl"] = None
    if "df_clean" not in st.session_state:
        st.session_state["df_clean"] = None
    if "scale_df_used" not in st.session_state:
        st.session_state["scale_df_used"] = None
    if "fp_df_used" not in st.session_state:
        st.session_state["fp_df_used"] = None

    colA, colB = st.columns(2)
    with colA:
        data_file = st.file_uploader("📥 Fichier résultats", type=["xlsx"], accept_multiple_files=False, key="ranking_data")
    with colB:
        scale_file = st.file_uploader("📥 Point scale", type=["xlsx"], accept_multiple_files=False, key="ranking_scale")

    st.divider()
    st.subheader("Date de référence pour le coeff temps")
    date_ref_input = st.date_input("Date de référence", value=dt.date.today(), key="ranking_date")

    st.divider()
    st.subheader("Points de finale (Rank → Points)")
    fp_upload = st.file_uploader(
        "Optionnel: importer une table points finale (csv/xlsx)",
        type=["csv", "xlsx"],
        accept_multiple_files=False,
        key="ranking_fp_upload",
    )

    if fp_upload is None:
        final_points_df = default_final_points()
    else:
        ext = fp_upload.name.split(".")[-1].lower()
        if ext == "csv":
            final_points_df = pd.read_csv(fp_upload)
        else:
            final_points_df = pd.read_excel(fp_upload, engine=EXCEL_ENGINE)

    final_points_df = st.data_editor(final_points_df, num_rows="dynamic", key="editor_fp", use_container_width=True)

    st.divider()
    st.subheader("Coeff compétitions (match → coeff)")

    def default_coeff_table():
        return pd.DataFrame({
            "MATCH": ["EN", "LOC", "GP", "INT", "CDF", "JECH", "ECH", "JWCH", "WCH", "JWC", "WC", "WCF"],
            "COEFF": [2, 0.5, 3, 2, 1, 3, 3, 5, 5, 4, 4, 5],
        })

    coeff_upload = st.file_uploader(
        "Optionnel : importer une table coeff (csv/xlsx)",
        type=["csv", "xlsx"],
        accept_multiple_files=False,
        key="ranking_coeff_upload",
    )

    if coeff_upload is None:
        coeff_df = default_coeff_table()
    else:
        ext = coeff_upload.name.split(".")[-1].lower()
        coeff_df = pd.read_csv(coeff_upload) if ext == "csv" else pd.read_excel(coeff_upload, engine=EXCEL_ENGINE)

    coeff_df = st.data_editor(coeff_df, num_rows="dynamic", key="editor_coeff", use_container_width=True)

    st.subheader("Barème (Point scale)")
    if scale_file is not None:
        scale_df = parse_scale_from_excel(scale_file.getvalue())
        if scale_df.empty:
            st.warning("Je n’ai pas trouvé les blocs (ex: '50m Women', '10m Junior Men', etc.).")
    else:
        scale_df = pd.DataFrame({"ScaleKey": [], "MinScore": [], "Points": []})

    scale_df = st.data_editor(scale_df, num_rows="dynamic", key="editor_scale", use_container_width=True)

    st.divider()
    st.subheader("Athlètes à exclure")
    remove_text = st.text_area(
        "1 nom par ligne",
        value="BAILLY Nathan\nFORET Alexandre\nGERMOND Etienne\nMOMBERT Claire",
        height=110,
        key="ranking_remove_text",
    )
    athletes_to_remove = [x.strip() for x in remove_text.splitlines() if x.strip()]

    run = st.button("Calculer le tableau final", type="primary", key="ranking_run")

    if run:
        if data_file is None:
            st.error("Upload le fichier Excel résultats (FFTir) d’abord.")
            return
        if scale_df.empty:
            st.error("Il me faut un barème (point scale).")
            return

        scale_df_c = scale_df.copy()
        scale_df_c["MinScore"] = pd.to_numeric(scale_df_c["MinScore"], errors="coerce")
        scale_df_c["Points"] = pd.to_numeric(scale_df_c["Points"], errors="coerce")
        scale_df_c = scale_df_c.dropna(subset=["ScaleKey", "MinScore", "Points"])

        fp = final_points_df.copy()
        fp_map = dict(zip(pd.to_numeric(fp["Rank"]).astype(int), pd.to_numeric(fp["FinalPoints"]).astype(float)))
        fp = round_existing_columns(fp, ["Rank", "FinalPoints"], 1)

        df = extract_workbook(data_file.getvalue())
        df = df[df["Score"].fillna(0) > 0].copy()
        df = df[df["Score"].notna()].copy()
        df = derive_sex_distance(df)
        df = add_comp_coeff(df, coeff_df)
        df = fix_names(df)

        if athletes_to_remove:
            df = df[~df["Athlete"].isin(athletes_to_remove)].copy()

        df = add_date_coeff(df, date_ref_input)
        df["ScaleKey_S"] = df.apply(lambda r: scale_key(r["Distance"], r["Sexe"], "S"), axis=1)
        df["ScaleKey_J"] = df.apply(lambda r: scale_key(r["Distance"], r["Sexe"], r["Category"]), axis=1)

        df["Perf_S"] = df.apply(lambda r: score_to_points(r["Score"], scale_df_c, r["ScaleKey_S"]), axis=1)
        df["Perf_J"] = df.apply(lambda r: score_to_points(r["Score"], scale_df_c, r["ScaleKey_J"]), axis=1)
        df["Finale_score"] = df["Rank"].apply(lambda x: fp_map.get(int(x), np.nan) if pd.notna(x) else np.nan)

        df["Total_Score_J"] = df["Perf_J"] + df["Finale_score"].fillna(0)
        df["Total_Score_S"] = df["Perf_S"] + df["Finale_score"].fillna(0)
        df["%J"] = (df["Total_Score_J"] / DENOM) * 100.0
        df["%S"] = (df["Total_Score_S"] / DENOM) * 100.0

        df = df[df["Distance"].notna()].copy()
        final_tbl = make_final_table(df, date_ref_input)

        st.session_state["final_tbl"] = final_tbl
        st.session_state["df_clean"] = df
        st.session_state["scale_df_used"] = scale_df_c
        st.session_state["fp_df_used"] = fp

    if st.session_state["final_tbl"] is not None:
        df = st.session_state["df_clean"]
        final_tbl = st.session_state["final_tbl"]
        scale_df_used = st.session_state["scale_df_used"]
        fp_used = st.session_state["fp_df_used"]

        tab1, tab2 = st.tabs(["Données clean (long)", "Tableau final"])

        with tab1:
            st.dataframe(df, use_container_width=True, height=520)
            st.subheader("Détail calcul %S compet / %J compet")

            athlete_list = sorted(df["Athlete"].dropna().unique())
            athlete_debug = st.selectbox("Choisir une athlète", athlete_list, key="athlete_sel")

            debug_df = df[df["Athlete"] == athlete_debug].copy()
            debug_df["is_12m"] = debug_df["Date"].apply(lambda d: within_last_12_months(d, date_ref_input))
            debug_df["Poids_%S_compet"] = debug_df["Coeff"] * debug_df["date_coeff"]
            debug_df["Contribution_%S_compet"] = debug_df["%S"] * debug_df["Poids_%S_compet"]
            debug_df["Poids_%J_compet"] = debug_df["Coeff"] * debug_df["date_coeff"]
            debug_df["Contribution_%J_compet"] = debug_df["%J"] * debug_df["Poids_%J_compet"]

            debug_df = round_existing_columns(
                debug_df,
                [
                    "Score", "Rank", "Perf_S", "Perf_J", "Finale_score", "%S", "%J",
                    "Coeff", "date_coeff", "Poids_%S_compet", "Contribution_%S_compet",
                    "Poids_%J_compet", "Contribution_%J_compet",
                ],
                1,
            )

            st.dataframe(
                debug_df[[
                    "Athlete", "Distance", "Competition", "Date", "is_12m", "Score",
                    "Perf_S", "Perf_J", "Finale_score", "%S", "%J",
                    "Coeff", "date_coeff",
                    "Poids_%S_compet", "Contribution_%S_compet",
                    "Poids_%J_compet", "Contribution_%J_compet",
                ]],
                use_container_width=True,
                height=320,
            )

            if not debug_df.empty:
                def get_res(num_col, den_col):
                    n, d = debug_df[num_col].sum(), debug_df[den_col].sum()
                    return n, d, (n / d if d != 0 else np.nan)

                ns, ds, rs = get_res("Contribution_%S_compet", "Poids_%S_compet")
                nj, dj, rj = get_res("Contribution_%J_compet", "Poids_%J_compet")

                c1, c2 = st.columns(2)
                with c1:
                    st.write("**Formule %S compet**")
                    st.caption(f"Num (Σ %S * coeffs): {ns:.1f} / Den (Σ coeffs): {ds:.1f} = **{rs:.1f}**")
                with c2:
                    st.write("**Formule %J compet**")
                    st.caption(f"Num (Σ %J * coeffs): {nj:.1f} / Den (Σ coeffs): {dj:.1f} = **{rj:.1f}**")

        with tab2:
            st.dataframe(final_tbl, use_container_width=True, height=520)
            st.subheader("Exports")
            c1, c2 = st.columns(2)
            with c1:
                st.download_button(
                    "Télécharger CSV",
                    data=final_tbl.to_csv(index=False).encode("utf-8"),
                    file_name="tableau_final.csv",
                    mime="text/csv",
                    key="ranking_csv",
                )
            with c2:
                out = io.BytesIO()
                with pd.ExcelWriter(out, engine="openpyxl") as writer:
                    final_tbl.to_excel(writer, index=False, sheet_name="Final")
                    df.to_excel(writer, index=False, sheet_name="Clean_Long")
                    scale_df_used.to_excel(writer, index=False, sheet_name="Scale")
                    fp_used.to_excel(writer, index=False, sheet_name="FinalPoints")

                    ws = writer.book["Final"]
                    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    orange = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
                    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

                    headers = [c.value for c in ws[1]]
                    for name in ["%S", "%J"]:
                        if name in headers:
                            col_letter = ws.cell(row=1, column=headers.index(name) + 1).column_letter
                            cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
                            ws.conditional_formatting.add(cell_range, CellIsRule(operator="lessThan", formula=["30"], fill=red))
                            ws.conditional_formatting.add(cell_range, CellIsRule(operator="between", formula=["30", "70"], fill=orange))
                            ws.conditional_formatting.add(cell_range, CellIsRule(operator="greaterThan", formula=["70"], fill=green))

                st.download_button(
                    "Télécharger XLSX",
                    data=out.getvalue(),
                    file_name="tableau_final.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="ranking_xlsx",
                )


# =========================================================
# APP NAVIGATION
# =========================================================
with st.sidebar:
    st.markdown("---")
    page = st.radio(
        "Navigation",
        ["Page 1 — Analyse CDF", "Page 2 — Ranking App"],
        key="main_navigation",
    )

if page == "Page 1 — Analyse CDF":
    render_page_cdf()
else:
    render_page_ranking()
